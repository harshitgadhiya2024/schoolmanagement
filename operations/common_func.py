"""
Maintain common function code like we can use all function in every where
"""

import logging
from datetime import datetime
from constant import constant_data
import random
import phonenumbers
import pandas as pd
import json
import os
import pycountry
import csv
import openpyxl

from operations.mongo_connection import data_added


def logger_con(app):
    """
    Configure logger with filename

    :param app: app-name
    :return:
    """
    try:
        server_file_name = constant_data.get("log_file_name", constant_data.get("log_file_name", "server.log"))
        logging.basicConfig(filename=server_file_name, level=logging.DEBUG)

    except Exception as e:
        app.logger.error(f"Error when connecting logger: {e}")


def get_timestamp(app):
    """
    Get current time

    :param app: app-name
    :return: current_date
    """
    try:
        current_datetime = datetime.now()
        formatted_datetime = current_datetime.strftime("%m-%d-%Y %H:%M:%S")

        return formatted_datetime

    except Exception as e:
        app.logger.debug(f"Error in get timestamp: {e}")


def get_error_msg(app, e):
    """
    Get error response with formatting

    :param app: app-name
    :param e: error
    :return: error-response
    """
    try:
        response_data_msg = {
            "data": "null",
            "message": f"Server Not Responding Error {e}",
            "status": "FORBIDDEN",
            "statusCode": 403,
            "timestamp": get_timestamp(app)
        }

        return response_data_msg

    except Exception as e:
        app.logger.debug(f"Error in get response msg: {e}")

def get_response_msg(app,status, statuscode, message, data):
    """
    Get success response with formatting

    :param app: app-name
    :param status: status-name
    :param statuscode: status-code like 200,300
    :param message: status message
    :param data: response data
    :return: formatted response
    """

    try:
        response_data_msg = {
            "timestamp": get_timestamp(app),
            "status": status,
            "statusCode": statuscode,
            "message": message,
            "data": data
        }

        return response_data_msg

    except Exception as e:
        app.logger.debug(f"Error in get response msg: {e}")

def get_unique_student_id(app, all_student_id):
    """
    Get unique student id

    :param app: app-name
    :param all_student_id: registered all student ids
    :return: unique student id
    """
    try:
        flag = True
        while flag:
            student_id = random.randint(100000000000, 999999999999)
            if str(student_id) not in all_student_id:
                flag = False

        return str(student_id)

    except Exception as e:
        app.logger.debug(f"Error in get unique student id: {e}")


def get_unique_department_id(app, all_deparment_id):
    """
    Get unique department id

    :param app: app-name
    :param all_deparment_id: registered all department ids
    :return: unique department id
    """
    try:
        flag = True
        while flag:
            depart_id = random.randint(100000, 999999)
            if str(depart_id) not in all_deparment_id:
                flag = False

        return str(depart_id)

    except Exception as e:
        app.logger.debug(f"Error in get unique department id: {e}")


def get_unique_subject_id(app, all_subject_id):
    """
    Get unique department id

    :param app: app-name
    :param all_subject_id: registered all subject ids
    :return: unique subject id
    """
    try:
        flag = True
        while flag:
            sub_id = random.randint(1000, 9999)
            if str(sub_id) not in all_subject_id:
                flag = False

        return str(sub_id)

    except Exception as e:
        app.logger.debug(f"Error in get unique subject id: {e}")


def get_unique_teacher_id(app, all_teacher_id):
    """
    Get unique teacher id

    :param app: app-name
    :param all_teacher_id: registered all teacher ids
    :return: unique teacher id
    """
    try:
        flag = True
        while flag:
            teacher_id = random.randint(10000000000000, 99999999999999)
            if str(teacher_id) not in all_teacher_id:
                flag = False

        return str(teacher_id)

    except Exception as e:
        app.logger.debug(f"Error in get unique teacher: {e}")


def get_unique_admin_id(app, all_admin_id):
    """
    Get unique teacher id

    :param app: app-name
    :param all_teacher_id: registered all teacher ids
    :return: unique teacher id
    """
    try:
        flag = True
        while flag:
            admin_id = random.randint(1000000000000000, 9999999999999999)
            if str(admin_id) not in all_admin_id:
                flag = False

        return str(admin_id)

    except Exception as e:
        app.logger.debug(f"Error in get unique admin: {e}")


def password_validation(app, password):
    """
    Check password is validate or not

    :param app: app-name
    :param password: user password
    :return: True or False
    """
    try:
        capital = False
        number = False
        special_char = False
        special_char_list = ["@", "#", "$", "-", "_", "*"]
        for char in password:
            if char.isupper():
                capital = True
            try:
                char = int(char)
                number = True
            except:
                pass

            if char in special_char_list:
                special_char = True

        if capital and number and special_char:
            matching = True
        else:
            matching = False

        return matching

    except Exception as e:
        app.logger.debug(f"Error in validate password: {e}")


def validate_phone_number(app, phone_number):
    """
    Check phone number is validate or not

    :param app: app-name
    :param phone_number: user phone-number
    :return: True or False
    """
    try:
        # Parse the phone number
        parsed_number = phonenumbers.parse(phone_number, None)

        # Check if the number is valid
        is_valid = phonenumbers.is_valid_number(parsed_number)

        if is_valid:
            return "valid number"
        else:
            return "invalid number"

    except Exception as e:
        app.logger.debug(f"Error in validate password: {e}")


def get_admin_data(app, client, db_name, coll_name):
    """
    get all data from student, admin and teacher database table

    :param app: app-name
    :param client: mongo client
    :param db_name: database-name
    :param coll_name: collection-name
    :return: database_unique_keys and database_userdata
    """
    try:
        db = client[db_name]
        coll = db[coll_name]
        res = coll.find({})
        keys = []
        values = []
        for each_res in res:
            all_keys = list(each_res.keys())
            if all_keys[1:] not in keys:
                keys.append(all_keys[1:])

            all_values = list(each_res.values())
            values.append(all_values[1:])

        return keys, values

    except Exception as e:
        app.logger.debug(f"Error in get data from admin database: {e}")

def get_student_data(app, client, db_name, coll_name, condition_dict):
    """
    get all data from student, admin and teacher database table

    :param app: app-name
    :param client: mongo client
    :param db_name: database-name
    :param coll_name: collection-name
    :return: database_unique_keys and database_userdata
    """
    try:
        db = client[db_name]
        coll = db[coll_name]
        res = coll.find(condition_dict)
        keys = []
        values = []
        for each_res in res:
            all_keys = list(each_res.keys())
            if all_keys[1:] not in keys:
                keys.append(all_keys[1:])

            all_values = list(each_res.values())
            values.append(all_values[1:])

        return keys, values

    except Exception as e:
        app.logger.debug(f"Error in get data from admin database: {e}")

def get_profile_data(app, client, db_name, type):
    """
    get all data from student, admin and teacher database table

    :param app: app-name
    :param client: mongo client
    :param db_name: database-name
    :param coll_name: collection-name
    :return: database_unique_keys and database_userdata
    """
    try:
        if type=="admin":
            coll_name = "admin_data"
        elif type=="teacher":
            coll_name = "teacher_data"
        else:
            coll_name = "students_data"

        db = client[db_name]
        coll = db[coll_name]
        res = coll.find({})
        res = list(res)
        return res[0]

    except Exception as e:
        app.logger.debug(f"Error in get profile data for database: {e}")


def delete_panel_data(app, client, db_name, coll_name, delete_dict):
    """
    delete data from database

    :param app: app-name
    :param client: mongo-client
    :param db_name: database-name
    :param coll_name: collection-name
    :param delete_dict: condition dict
    :return: status
    """

    try:
        type = delete_dict.get("type", "else")
        db = client[db_name]
        coll = db[coll_name]
        get_all_data = coll.find(delete_dict)
        if type == "student" or type == "teacher" or type == "admin":
            for delete_photo in get_all_data:
                photo_path = delete_photo["photo_link"]
                if photo_path != "static/assets/img/profiles/avatar-01.jpg":
                    if os.path.isfile(photo_path):
                        os.remove(photo_path)
            coll.delete_one(delete_dict)
            coll1 = db["login_mapping"]
            coll1.delete_one(delete_dict)
            if type == "student":
                coll_class = db["class_data"]
                coll_class.delete_one(delete_dict)
            elif type == "teacher":
                coll_class = db["subject_mapping"]
                coll_class.delete_one(delete_dict)
        else:
            coll.delete_one(delete_dict)

        return "data deleted"

    except Exception as e:
        app.logger.debug(f"Error in delete data from database: {e}")

def delete_teacher_panel_data(app, client, db_name, coll_name, delete_dict):
    """
    delete data from database

    :param app: app-name
    :param client: mongo-client
    :param db_name: database-name
    :param coll_name: collection-name
    :param delete_dict: condition dict
    :return: status
    """

    try:
        db = client[db_name]
        coll = db[coll_name]
        coll.delete_one(delete_dict)

        return "data deleted"

    except Exception as e:
        app.logger.debug(f"Error in delete data from database: {e}")


def delete_all_panel_data(app, client, db_name, coll_name, panel):
    """
    delete data from database

    :param app: app-name
    :param client: mongo-client
    :param db_name: database-name
    :param coll_name: collection-name
    :return: status
    """

    try:
        db = client[db_name]
        coll = db[coll_name]
        count = coll.count_documents({})
        print(f"count: {count}")
        if (panel == "student" or panel == "teacher" or panel == "admin") and count > 0:
            get_all_data = coll.find({})
            for delete_photo in get_all_data:
                photo_path = delete_photo["photo_link"]
                if photo_path != "static/assets/img/profiles/avatar-01.jpg":
                    if os.path.isfile(photo_path):
                        os.remove(photo_path)
            coll.delete_many({})
            coll1 = db["login_mapping"]
            coll1.delete_many({"type": panel})
            if panel == "student":
                coll_class = db["class_data"]
                coll_class.delete_many({"type": panel})
            elif panel=="teacher":
                coll_class = db["subject_mapping"]
                coll_class.delete_many({"type": panel})
            return True
        else:
            coll.delete_many({})
            return False

    except Exception as e:
        print(f"Error in delete data from database: {e}")
        app.logger.debug(f"Error in delete data from database: {e}")


def export_panel_data(app, database_data, panel, type):
    """
    export data for different format like csv, excel and csv

    :param app: app-name
    :param database_data: database data
    :param type: excel, csv, json
    :return: filename
    """

    try:
        if type == "excel":
            output_path = os.path.join(app.config["EXPORT_UPLOAD_FOLDER"], f"export_{panel}_excel.xlsx")
            df = pd.DataFrame(database_data)
            df.to_excel(output_path, index=False)
        elif type == "csv":
            output_path = os.path.join(app.config["EXPORT_UPLOAD_FOLDER"], f"export_{panel}_csv.csv")
            df = pd.DataFrame(database_data)
            df.to_csv(output_path, index=False)
        else:
            output_path = os.path.join(app.config["EXPORT_UPLOAD_FOLDER"], f"export_{panel}_json.json")
            with open(output_path, 'w') as json_file:
                json.dump(database_data, json_file, indent=2)

        return output_path

    except Exception as e:
        app.logger.debug(f"Error in export data from database: {e}")


def export_student_panel_data(app, database_data, panel, type):
    """
    export data for different format like csv, excel and csv

    :param app: app-name
    :param database_data: database data
    :param type: excel, csv, json
    :return: filename
    """

    try:
        if type == "excel":
            output_path = os.path.join(app.config["EXPORT_UPLOAD_FOLDER"], f"export_student_{panel}_excel.xlsx")
            df = pd.DataFrame(database_data)
            df.to_excel(output_path, index=False)
        elif type == "csv":
            output_path = os.path.join(app.config["EXPORT_UPLOAD_FOLDER"], f"export_student_{panel}_csv.csv")
            df = pd.DataFrame(database_data)
            df.to_csv(output_path, index=False)
        else:
            output_path = os.path.join(app.config["EXPORT_UPLOAD_FOLDER"], f"export_student_{panel}_json.json")
            with open(output_path, 'w') as json_file:
                json.dump(database_data, json_file, indent=2)

        return output_path

    except Exception as e:
        app.logger.debug(f"Error in export data from database: {e}")

def export_teacher_panel_data(app, database_data, panel, type):
    """
    export data for different format like csv, excel and csv

    :param app: app-name
    :param database_data: database data
    :param type: excel, csv, json
    :return: filename
    """

    try:
        if type == "excel":
            output_path = os.path.join(app.config["EXPORT_UPLOAD_FOLDER"], f"export_teacher_{panel}_excel.xlsx")
            df = pd.DataFrame(database_data)
            df.to_excel(output_path, index=False)
        elif type == "csv":
            output_path = os.path.join(app.config["EXPORT_UPLOAD_FOLDER"], f"export_teacher_{panel}_csv.csv")
            df = pd.DataFrame(database_data)
            df.to_csv(output_path, index=False)
        else:
            output_path = os.path.join(app.config["EXPORT_UPLOAD_FOLDER"], f"export_teacher_{panel}_json.json")
            with open(output_path, 'w') as json_file:
                json.dump(database_data, json_file, indent=2)

        return output_path

    except Exception as e:
        app.logger.debug(f"Error in export data from database: {e}")


def search_panel_data(app, client, db_name, search_value, coll_name):
    """
    search data from the database based on panel and search value

    :param app: app-name
    :param panel: panel
    :param search_value: id, name, phone numer, email
    :return: data regarding search value
    """

    try:
        db = client[db_name]
        coll = db[coll_name]
        print(f"search_value: {search_value} and type: {type(search_value)}")
        key, value = search_value.split("|")
        if key == "class_id":
            key = "student_id"
        print(f"key: {key}, value: {value}")
        data_type_converters = {
            "admin_data": {
                "admin_id": int,
            },
            "teacher_data": {
                "teacher_id": int,
            },
            "department_data": {
                "department_id": int,
            },
            "subject_data": {
                "subject_id": int,
            },
            "default": {
                "student_id": int,
            }
        }

        converter = data_type_converters.get(coll_name, data_type_converters["default"])
        print(f"converter: {converter}")
        search_value = converter.get(key, str)(value.strip())
        print(f"search_value: {search_value} and type: {type(search_value)}")

        result = coll.find_one({key: search_value})
        print(f"result: {result}")
        return result

    except Exception as e:
        print(f"Exception in searching panel data: {e}")
        app.logger.debug(f"Error in search data from database: {e}")


def get_all_country_state_names(app):
    try:
        netherlands_subdivisions = pycountry.subdivisions.get(country_code='NL')
        all_state_names = [subdivision.name for subdivision in netherlands_subdivisions]
        country_codes = phonenumbers.COUNTRY_CODE_TO_REGION_CODE
        list_country_code = []
        for country_code, region_codes in country_codes.items():
            list_country_code.append("+" + str(country_code))

        return all_state_names, list_country_code

    except Exception as e:
        app.logger.debug(f"Error in get country, state or city from database: {e}")


def import_data_into_database(app, db, panel, reader_coll):
    try:
        print(f"Data: {reader_coll}")
        print(f"Panel: {panel}")

        common_dict = {
            "inserted_on": get_timestamp(app),
            "updated_on": get_timestamp(app)
        }

        panel_dict = {
            "admin": {"type": "admin"},
            "student": {"type": "student"},
            "teacher": {"type": "teacher"},
            "department": {"type": "department"},
            "subject": {"type": "subject"},
            "class": {"type": "class"},
        }

        register_dict = {}
        for key in ["username", "first_name", "last_name", "password", "gender", "contact_no",
                    "emergency_contact_no", "email", "address", "city", "state", "country"]:
            register_dict[key] = reader_coll.get(key, "NA")

        if panel == "admin":
            register_dict = {
                "admin_id": reader_coll.get("unique_admin_id", "admin_id"),
                "type": "admin",
            }
            admin_mapping_dict = {key: reader_coll[key] for key in ["admin_id", "username", "email", "password"]}
            admin_mapping_dict["type"] = "admin"
            [data_added(app, db, ["admin_data", "login_mapping"], [register_dict, admin_mapping_dict])]
        elif panel == "student":
            register_dict = {
                "student_id": reader_coll.get("unique_student_id", "student_id"),
                "dob": reader_coll.get("dob", "NA"),
                "admission_date": reader_coll.get("admission_date", "NA"),
                "classes": reader_coll.get("classes", ""),
                "department": reader_coll.get("department", "NA"),
                "batch_year": reader_coll.get("batch_year", "NA"),
                **panel_dict[panel],
                **common_dict
            }

            student_mapping_dict = {key: reader_coll.get(key, "") for key in
                                    ["student_id", "username", "email", "password"]}
            student_mapping_dict["type"] = "student"
            classes_mapping_dict = {"student_id": reader_coll.get("student_id"),
                                    "department": reader_coll.get("department", "NA"),
                                    "class_name": reader_coll.get("class_name", "NA"),
                                    **panel_dict[panel]}

            [data_added(app, db, coll_name, new_dict)for coll_name, new_dict in zip(["class_data", "students_data", "login_mapping"], [register_dict, classes_mapping_dict, student_mapping_dict])]
        elif panel == "teacher":
            register_dict = {
                "teacher_id": reader_coll.get("unique_teacher_id", "teacher_id"),
                "dob": reader_coll.get("dob", "NA"),
                "qualification": reader_coll.get("qualification", "NA"),
                "department": reader_coll.get("department", "NA"),
                "subject": reader_coll.get("subject", "NA"),
                "joining_date": reader_coll.get("joining_date", "NA"),
                **panel_dict[panel],
                **common_dict
            }

            teacher_mapping_dict = {key: reader_coll.get(key, "") for key in
                                    ["teacher_id", "username", "email", "password"]}
            teacher_mapping_dict["type"] = "teacher"
            subject_mapping_dict = {"teacher_id": reader_coll.get("teacher_id"),
                                    "department_name": reader_coll.get("department", "NA"),
                                    "subject": reader_coll.get("subject", "NA"),
                                    **panel_dict[panel]}

            [data_added(app, db, coll_name, new_dict)for coll_name, new_dict in zip(["subject_mapping", "teacher_data", "login_mapping"], [subject_mapping_dict, register_dict, teacher_mapping_dict])]
        elif panel == "department":
            register_dict = {
                "department_id": reader_coll.get("unique_department_id", "department_id"),
                "department_name": reader_coll.get("department_name", "NA"),
                "department_date": reader_coll.get("department_date", "NA"),
                "HOD_name": reader_coll.get("HOD_name", "NA"),
                **panel_dict[panel],
                **common_dict
            }

            data_added(app, db, "department_data", register_dict)
        elif panel == "subject":
            register_dict = {
                "subject_id": reader_coll.get("unique_subject_id", "subject_id"),
                "subject_name": reader_coll.get("subject_name", "NA"),
                "department_name": reader_coll.get("department_name", "NA"),
                "subject_start_date": reader_coll.get("subject_start_date", "NA"),
                **panel_dict[panel],
                **common_dict
            }
            data_added(app, db, "subject_data", register_dict)
        else:
            register_dict = {
                "subject_id": reader_coll.get("unique_subject_id", "subject_id"),
                "subject_name": reader_coll.get("subject_name", "NA"),
                "department_name": reader_coll.get("department_name", "NA"),
                "subject_start_date": reader_coll.get("subject_start_date", "NA"),
                **panel_dict[panel],
                **common_dict
            }
            data_added(app, db, "subject_data", register_dict)
        print("Successfully added data to database")
        app.logger.debug("Successfully added data to database")
        return True
    except Exception as e:
        print(f"Error in importing data to database: {e}")
        app.logger.debug(f"Error in importing data to database: {e}")
        return False


def file_check(app, file_extension, file_path):
    try:
        if file_extension == ".csv" or file_extension == ".xlsx" or file_extension == ".json":
            ## Converting data to json for multiple file types
            if file_extension == ".csv":
                with open(file_path) as f:
                    reader = csv.DictReader(f)
                    field_names = reader.fieldnames
                    print(f"Field names: {field_names}")
                    rows = list(reader)
                    reader_json = json.loads(json.dumps(rows))
                    print(f"Type of reader_json: {type(reader_json)}")
                    print("Data in json format: ", reader_json)
            elif file_extension == ".xlsx":
                wb = openpyxl.load_workbook(file_path)
                sheet = wb.active
                field_names = [cell.value for cell in sheet[1]]
                print(f"Field names: {field_names}")
                # Convert the DataFrame to JSON
                # Create an empty list to store the JSON data
                json_data = []

                # Iterate over the rows in the sheet
                for row in sheet.iter_rows(values_only=True):
                    # Create a dictionary for each row
                    row_dict = {}
                    for col_index, cell_value in enumerate(row):
                        # Use the header row as keys for each dictionary entry
                        header = sheet.cell(row=1, column=col_index + 1).value
                        row_dict[header] = cell_value
                    # Append the row dictionary to the json_data list
                    json_data.append(row_dict)

                # Convert the list of dictionaries to JSON
                reader_json = json.load(json.dumps(json_data))
                print(f"Type of reader_json: {type(reader_json)}")
                # Print the JSON data
                print(f"json_string: {reader_json}")
            else:
                with open(file_path) as f:
                    reader_json = json.load(f)
                field_names = list(reader_json[0].keys())
                print(f"Field names: {field_names}")
                print(f"Data is in json format: {reader_json}")

            ## Removing keys that are not to be used in import data
            remove_key_bool, reader_json = remove_unused_keys(app, reader_json, field_names)
            print()
            if remove_key_bool:
                return True, field_names, reader_json
            else:
                print("Unable to remove unused keys from json file")
                app.logger.debug("Unable to remove unused keys from json file")
                return False, {}, {}
        else:
            print("Please select correct file format (.csv, .xlsx or .json)")
            app.logger.debug("Please select correct file format (.csv, .xlsx or .json)")
            return False, {}, {}
    except Exception as e:
        print(f"Error in file check: {e}")
        app.logger.debug(f"Error in file check: {e}")
        return None, {}, {}


def remove_unused_keys(app, reader_json, field_names):
    try:
        ## Keys that are to be manipulated
        remove_keys = ["_id", "updated_on", "photo_link"]
        for remove_key in remove_keys:
            if remove_key in field_names:
                print("Removing key in reader json: ", remove_key)
                for reader_dict in reader_json:
                    ## Updating this key with new timestamp
                    if remove_key == "updated_on":
                        reader_dict[remove_key] = get_timestamp(app)
                        print("Updated timestamp")
                    ## Removing link to photo ID
                    elif remove_key == "photo_link":
                        reader_dict[remove_key] = "NA"
                    ## Deleting _id from import data
                    else:
                        del reader_dict[remove_key]
                        print("Removed key in reader json")
        print(f"Updated json : {reader_json}")
        return True, reader_json
    except Exception as e:
        print(f"Error in remove unused keys: {e}")
        app.logger.debug(f"Error in remove unused keys: {e}")
        return False, reader_json


def check_dirs(app, file_name, file):
    try:
        if not os.path.isdir(app.config['IMPORT_UPLOAD_FOLDER']):
            os.makedirs(app.config['IMPORT_UPLOAD_FOLDER'], exist_ok=True)
        if not os.path.exists(app.config['REJECTED_DATA_UPLOAD_FOLDER']):
            os.makedirs(app.config['REJECTED_DATA_UPLOAD_FOLDER'], exist_ok=True)
        file_path = os.path.join(app.config['IMPORT_UPLOAD_FOLDER'], file_name)
        file.save(file_path)
        file_extension = os.path.splitext(file_name)[1]
        return file_extension, file_path
    except Exception as e:
        print(f"Error in creating dirs: {e}")
        app.logger.debug(f"Error in creating dirs: {e}")
        return None, None


def create_query_list(app, panel_obj, reader_json, file_name):
    try:
        check_id_mapping = {"admin": "admin_id", "student": "student_id", "teacher": "teacher_id",
                            "department": "department_id", "subject": "subject_id", "class": "student_id"}
        used_panel = check_id_mapping[panel_obj]
        query_and_data_dict = {}
        rejected_data = []
        for json_place_value in range(len(reader_json)):
            json_dict = reader_json[json_place_value]
            print(f"Json group: {json_dict}")
            if used_panel in json_dict:
                print("In here")
                query = used_panel + "|" + str(json_dict[used_panel])
                print(f"Query key: {query} and type of query key: {type(query)}")
                json_dict[used_panel] = int(json_dict[used_panel])
                query_and_data_dict[query] = json_dict
            else:
                print("Adding data to rejected file")
                rejected_data.append(json_dict)
                del reader_json[json_place_value]
        if len(rejected_data) > 0:
            update_rejected_data_file(app, file_name, rejected_data)
        return query_and_data_dict, panel_obj, f"{panel_obj}_data"
    except Exception as e:
        print(f"Error in create query list: {e}")
        app.logger.debug(f"Error in create query list: {e}")
        return {}, None


def update_rejected_data_file(app, file_name, rejected_data):
    try:
        file_name_extension = get_timestamp(app)
        file_name_extension = file_name_extension.replace(":", "_").replace("-", "_").replace(" ", "_")
        updated_file_name = file_name + "_" + file_name_extension
        print(f"Updated file name: {updated_file_name}")
        rejected_data_path = os.path.join(app.config['REJECTED_DATA_UPLOAD_FOLDER'], file_name)
        with open(rejected_data_path, 'a') as f:
            json.dump(rejected_data, f)
    except Exception as e:
        print(f"Error in updating rejected data file: {e}")
        app.logger.debug(f"Error in updating rejected data file: {e}")

# def update_collection(app, collection, db, updated_reader_coll):
#     try:
#         result = db[collection].insert_one(updated_reader_coll)
#         if result:
#             app.logger.debug(f"Data inserted into collection: {collection}")
#             return True
#     except Exception as e:
#         app.logger.debug(f"Error in update collection: {e}")
#         print(f"Error in update collection: {e}")
#         return False
#
#
# def check_for_files(app, collection, db, reader_json):
#     try:
#         count = db[collection].count_documents({})
#         default_keys = []
#         if count > 0:
#             print('The collection has documents.')
#             first_record = db[collection].find_one({})
#             print(f"First record: {first_record}")
#             # Get the keys of the first record
#             default_keys = list(first_record.keys())
#             print(f"Keys: {default_keys}")
#             remove_keys = ["_id", "inserted_on", "updated_on", "photo_link"]
#             for remove_key in remove_keys:
#                 if remove_key in default_keys:
#                     default_keys.remove(remove_key)
#             print("Keys in json format updated: ", reader_json)
#             print("Updated keys: ", default_keys)
#             app.logger.debug("Documents present in collection")
#         else:
#             print('The collection is empty.')
#             app.logger.debug("Documents not present in collection")
#         return count, default_keys
#     except Exception as e:
#         app.logger.debug(f"Error in check for files: {e}")
#         print(f"Error in check for files: {e}")
#         return 0, []
